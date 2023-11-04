import random
import string
import logging
from logging.handlers import RotatingFileHandler
from openpyxl import load_workbook
import sys
import argparse

__author__ = "Ali Gurbuz"

def generate_password(user):
    special_chars = "!%&(),._-=^#!%&(),._-=^#"
    logger.debug("generated password")
    return f'{user[0]}{random.choice(special_chars)}{user[1]}' \
           f'{random.choice(special_chars)}{user[2]}'

def get_user():
    for row in ws.iter_rows(min_row=2):
        klasse = str(row[0].value).lower()
        raum = str(row[1].value)
        kv = str(row[2].value)
        yield klasse, raum, kv


def generate_scripts():
    with open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\create_class.sh", "w") as file:
        logger.debug("opened file " + file.name)
        print("#!/bin/bash \n", file=file)
    with open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\delete_class.sh", "w") as file:
        logger.debug("opened file " + file.name)
        print("#!/bin/bash  \n", file=file)
    open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\passwords_class", "w").close()
    create_user_entry(("lehrer",), ''.join(random.choice(string.ascii_letters) for _ in range(10)))
    create_user_entry(("seminar",), ''.join(random.choice(string.ascii_letters) for _ in range(10)))
    for user in get_user():
        pw = generate_password(user)
        create_user_entry(user, pw)




def create_user_entry(user, pwd):
    useradd(user, pwd)
    userdel(user)
    addpasswd(user, pwd)


def useradd(user, pwd):
    create = f'useradd -d /home/klassen/{"k" if user[0][0].isdigit() else ""}{user[0]} -c "{user[0]}" -m ' \
             f'-g cdrom,plugdev,sambashare -s /bin/bash {user[0]} && ' \
             f'echo {user[0]}:\"{pwd}\" | chpasswd'
    with open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\create_class.sh", "a") as file:
        logger.debug("opened file " + file.name)
        print(create, file=file)
        logger.info("wrote useradd into " + file.name)


def addpasswd(user, pwd):
    with open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\passwords_class", "a") as file:
        logger.debug("opened file " + file.name)
        print(user[0], pwd, file=file, sep=":")
        logger.info("wrote password into file for user in " + file.name)

def userdel(user):
    with open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\delete_class.sh", "a",
              encoding="utf-8") as file:
        file.write(f"userdel {user[0]} && rm -rf /home/klassen/{user[0]} \n")
        print(user[0])

if '__main__' == __name__:
    parser = argparse.ArgumentParser()
    output_group = parser.add_mutually_exclusive_group()
    output_group.add_argument("-v", "--verbose", help="increase output verbosity", action="store_true")
    output_group.add_argument("-q", "--quiet", help="decrease output verbosity", action="store_true")
    parser.add_argument("file", help="file with the userdata")
    args = parser.parse_args()

    logger = logging.getLogger()

    formatter = logging.Formatter("%(asctime)s; %(levelname)s; %(message)s",
                                  "%Y-%m-%d %H:%M:%S")

    rotating_file_handler = RotatingFileHandler(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\create_class.log", maxBytes=10000, backupCount=5)
    rotating_file_handler.setFormatter(formatter)
    logger.addHandler(rotating_file_handler)

    stream_handler = logging.StreamHandler(sys.stderr)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)


    if args.verbose:
        stream_handler.setLevel(logging.DEBUG)
        logger.setLevel(logging.DEBUG)
    elif args.quiet:
        stream_handler.setLevel(logging.CRITICAL)
        logger.setLevel(logging.Critical)
    else:
        logger.setLevel(logging.INFO)
        stream_handler.setLevel(logging.INFO)

    try:
        wb = load_workbook(args.file, read_only=True)
        ws = wb[wb.sheetnames[0]]
        generate_scripts()
    except FileNotFoundError:
        logger.critical("couldnt find file")



