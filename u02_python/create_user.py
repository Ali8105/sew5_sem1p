import argparse
import logging
import random
import re
import string
import unicodedata
from collections import namedtuple
from logging.handlers import RotatingFileHandler

from openpyxl import load_workbook


# toDo: Excel-Liste
#       unique Passwort


__author__ = "Ali Gurbuz"

def generate_password():
    """Generiert Passwort fur Benutzer"""
    pas_char = string.ascii_letters + "!%&(),._-=^#!%&(),._-=^#"
    logger.debug("generate password")
    return ''.join(random.choice(pas_char) for _ in range(10))


def generate_scripts():
    """Generiert die Skripts und iteriert uber die Benutzer"""
    with open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\script_user.sh", "w") as file:
        logger.debug("opened file " + file.name)
        print("#!/bin/bash \n", file=file)
    with open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\delete_user.sh", "w") as file:
        logger.debug("opened file " + file.name)
        print("#!/bin/bash  \n", file=file)
    open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\passwords_users", "w").close()
    # fur mehrere selbe Loginnamen
    users = dict()
    for user in get_user():
        login_name = user.nname
        counter = 1
        while login_name in users:
            # Entfernt alle Zahlen im login_name
            login_name = re.sub(r"(\d+)", "", login_name)
            login_name += str(counter)
            counter += 1
        users[login_name] = login_name
        pw = generate_password()
        user = user._replace(login_name=login_name)
        create_user_entry(user, pw)


def create_user_entry(user, pwd):
    """
    Schreibt eine Zeile im create- und delete script und in passwords_users fur den jeweiligen User
    :param user: USer
    :param pwd: Passwort
    """

    useradd(user,pwd)
    userdel(user)
    addpasswd(user, pwd)

def remove_accent(txt):
    """
    Normaliziert txt
    :param txt = Namen
    """
    norm_txt = unicodedata.normalize('NFD',txt)
    shaved = ''.join(c for c in norm_txt
                     if not unicodedata.combining(c))
    erg = unicodedata.normalize('NFC', shaved)
    return erg


def get_user():
    """Generator fur User"""
    User = namedtuple("User", "vname nname group u_class login_name")
    for row in ws.iter_rows(min_row=2):
        firstname = remove_accent(str(row[0].value).lower())
        lastname = remove_accent(str(row[1].value).lower())
        group = str(row[2].value)
        u_class = str(row[3].value)
        user = User(firstname, lastname, group, u_class, "")
        yield user



def useradd(user, pwd):
    """Useradd in script_user.sh"""

    # -d home Directory
    # -c Kommentar | Full Name
    # -m Benutzerverzeichnis
    # -g Hauptbenutzer falls Gruppe Stundets unter Gruppe Klasse
    # -s Standart Shell
    create = f'useradd -d "/home/{user.login_name}" -c "{user.vname + " " + user.nname}" -m ' \
             f'-g {user.group}{"," + user.u_class if user.group == "student" else ""} -s "/bin/bash {user.login_name}" && ' \
             f'echo {user.login_name}:\"{pwd}\" | chpasswd'

    with open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\script_user.sh", "a", encoding="utf-8") as file:
        logger.debug("opened file " + file.name)
        print(create, file=file)
        logger.info("wrote useradd into " + file.name)


def userdel(user):
    """Delete Script fur delete_user.sh"""
    with open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\delete_user.sh", "a") as file:
        logger.debug("opened file " + file.name)
        delete = f'userdel {user.login_name} && rm -rf /home/klassen/{user.login_name}'
        file.write(delete)
        logger.info("wrote userdel into " + file.name)

passworter = list()
def addpasswd(user, pwd):
    """Schreibt User und Passwort in passwords_user"""
    with open(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\passwords_users", "a") as file:
        logger.debug("opened file " + file.name)
        print(user.login_name, pwd, file=file, sep=":")
        logger.info("wrote password into file for user in " + file.name)
        while passworter.__contains__()



if '__main__' == __name__:
    parser = argparse.ArgumentParser()
    output_group = parser.add_mutually_exclusive_group()
    output_group.add_argument("-v", "--verbose", help="increase output verbosity", action="store_true")
    output_group.add_argument("-q", "--quiet", help="decrease output verbosity", action="store_true")
    parser.add_argument("file", help="file with the userdata")
    args = parser.parse_args()

    logger = logging.getLogger()

    formatter = logging.Formatter("%(asctime)s; %(levelname)s; %(message)s",
                                  "%Y-%m-%d %H:%M:%S")

    rotating_file_handler = RotatingFileHandler(r"C:\Users\aligr\Desktop\Schule\5CN\SEW\sew5_sem1p\Ressources\create_user.log", maxBytes=10000, backupCount=5)
    rotating_file_handler.setFormatter(formatter)
    logger.addHandler(rotating_file_handler)


    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)


    logger.setLevel(logging.DEBUG)

    if args.verbose:
        stream_handler.setLevel(logging.DEBUG)
    elif args.quiet:
        stream_handler.setLevel(logging.CRITICAL)
    else:
        stream_handler.setLevel(logging.INFO)

    try:
        wb = load_workbook(args.file, read_only=True)
        ws = wb[wb.sheetnames[0]]
        generate_scripts()
    except FileNotFoundError:
        logger.critical("couldnt find file")