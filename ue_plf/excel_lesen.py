import random
import string

import openpyxl
from openpyxl import load_workbook


def read_excel():
    wb = load_workbook(r"res\file.xlsx")
    ws = wb[wb.sheetnames[0]]

    for row in ws.iter_rows(values_only=True):
        create_user(row[0],row[1])


def create_user(name, klasse):
    rand_characters = string.ascii_letters + string.digits + string.punctuation
    passwort = ''.join(random.choice(rand_characters) for _ in range(random.randint(6,7)))
    print(passwort)
    with open(r"res/skript.ps", "a", encoding="utf-8") as script:
        script.write(f"New-LocalUser -Name {name} -Password {passwort} BlahBLAH \n")


if __name__ == "__main__":
    read_excel()